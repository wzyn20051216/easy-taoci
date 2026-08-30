"""导师追踪工作簿的备份与幂等同步。"""

from __future__ import annotations

import shutil
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from .core import read_jsonl


HEADERS = [
    "序号", "大学", "学院", "老师姓名", "研究方向", "是否已发送", "是否收到回复",
    "邮箱", "所属团队", "草稿状态", "来源URL", "匹配分", "任务ID", "更新时间",
]


def _load_openpyxl():
    """延迟加载可选依赖，保持基础安装轻量。"""
    try:
        import openpyxl
    except ImportError as exc:
        raise RuntimeError("同步 XLSX 需要安装：python -m pip install -e \".[xlsx]\"") from exc
    return openpyxl


def _headers(ws: Any) -> dict[str, int]:
    """补齐列并返回列名映射。"""
    if ws.max_row == 1 and ws.max_column == 1 and ws.cell(row=1, column=1).value is None:
        existing: list[str] = []
    else:
        existing = [str(cell.value) if cell.value is not None else "" for cell in ws[1]]
    for header in HEADERS:
        if header not in existing:
            ws.cell(row=1, column=len(existing) + 1, value=header)
            existing.append(header)
    return {
        str(ws.cell(row=1, column=column).value): column
        for column in range(1, ws.max_column + 1)
        if ws.cell(row=1, column=column).value is not None
    }


def _copy_style(ws: Any, row: int) -> None:
    """新增行时沿用上一行样式，不改动已有样式。"""
    if row <= 2:
        return
    for column in range(1, ws.max_column + 1):
        source = ws.cell(row=row - 1, column=column)
        target = ws.cell(row=row, column=column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format


def _upsert(ws: Any, draft: dict[str, Any], saved: bool, now: str) -> None:
    """按任务 ID 优先、学校姓名邮箱次优的方式更新一位导师。"""
    columns = _headers(ws)
    teacher = draft["teacher"]
    target_row = None
    for row in range(2, ws.max_row + 1):
        existing_task = str(ws.cell(row=row, column=columns["任务ID"]).value or "")
        existing_name = str(ws.cell(row=row, column=columns["老师姓名"]).value or "")
        existing_school = str(ws.cell(row=row, column=columns["大学"]).value or "")
        existing_email = str(ws.cell(row=row, column=columns["邮箱"]).value or "")
        if existing_task == draft["task_id"] or (
            existing_name == teacher["name"]
            and existing_school == teacher["university"]
            and existing_email == draft["recipient"]
        ):
            target_row = row
            break
    if target_row is None:
        target_row = ws.max_row + 1
        _copy_style(ws, target_row)
        ws.cell(row=target_row, column=columns["序号"], value=target_row - 1)

    sent_value = str(ws.cell(row=target_row, column=columns["是否已发送"]).value or "否")
    values = {
        "大学": teacher["university"],
        "学院": teacher["college"],
        "老师姓名": teacher["name"],
        "研究方向": teacher.get("research_focus", ""),
        "是否已发送": "是" if sent_value == "是" else "否",
        "是否收到回复": str(ws.cell(row=target_row, column=columns["是否收到回复"]).value or "否"),
        "邮箱": draft["recipient"],
        "草稿状态": "已存网易草稿" if saved else "已生成草稿",
        "来源URL": teacher.get("faculty_url", ""),
        "匹配分": teacher.get("match_score", ""),
        "任务ID": draft["task_id"],
        "更新时间": now,
    }
    for key, value in values.items():
        ws.cell(row=target_row, column=columns[key], value=value)


def sync_workbook(workbook_path: Path, drafts_path: Path, state_path: Path) -> Path:
    """备份并同步总表和学校分表，返回备份路径。"""
    openpyxl = _load_openpyxl()
    if not workbook_path.is_file():
        raise FileNotFoundError(f"工作簿不存在：{workbook_path}")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = workbook_path.with_name(f"{workbook_path.stem}_backup_{timestamp}{workbook_path.suffix}")
    shutil.copy2(workbook_path, backup)

    drafts = read_jsonl(drafts_path)
    states = read_jsonl(state_path) if state_path.is_file() else []
    saved_ids = {str(item.get("task_id")) for item in states if item.get("status") == "saved"}
    workbook = openpyxl.load_workbook(workbook_path)
    if "全部汇总" not in workbook.sheetnames:
        workbook.create_sheet("全部汇总", 0)
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for draft in drafts:
        school = str(draft["teacher"]["university"])
        if school not in workbook.sheetnames:
            workbook.create_sheet(school)
        saved = draft["task_id"] in saved_ids
        _upsert(workbook["全部汇总"], draft, saved, now)
        _upsert(workbook[school], draft, saved, now)

    workbook.save(workbook_path)
    return backup
