"""XLSX 追踪表同步回归测试。"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover - 可选依赖缺失时跳过
    openpyxl = None

from easy_taoci.core import write_jsonl
from easy_taoci.workbook import sync_workbook


@unittest.skipIf(openpyxl is None, "未安装 openpyxl")
class WorkbookTests(unittest.TestCase):
    """确保草稿状态与发送状态不会混淆。"""

    def test_saved_draft_remains_unsent_and_syncs_school_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "tracker.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "全部汇总"
            workbook.save(workbook_path)

            drafts_path = root / "drafts.jsonl"
            state_path = root / "state.jsonl"
            draft = {
                "task_id": "a" * 24,
                "recipient": "teacher@example.edu",
                "subject": "测试主题",
                "body": "测试正文",
                "match_paragraph": "测试匹配段",
                "attachments": [],
                "teacher": {
                    "university": "测试大学",
                    "college": "信息学院",
                    "name": "陈测试",
                    "research_focus": "智能感知",
                    "faculty_url": "https://faculty.example.edu/teacher",
                    "match_score": "88",
                },
            }
            write_jsonl(drafts_path, [draft])
            write_jsonl(state_path, [{"task_id": draft["task_id"], "status": "saved", "provider": "gmail"}])

            backup = sync_workbook(workbook_path, drafts_path, state_path)
            self.assertTrue(backup.is_file())
            result = openpyxl.load_workbook(workbook_path, data_only=True)
            self.assertIn("测试大学", result.sheetnames)
            for sheet_name in ("全部汇总", "测试大学"):
                sheet = result[sheet_name]
                headers = {cell.value: cell.column for cell in sheet[1]}
                self.assertEqual("否", sheet.cell(row=2, column=headers["是否已发送"]).value)
                self.assertEqual("已存Gmail草稿", sheet.cell(row=2, column=headers["草稿状态"]).value)
                self.assertEqual(draft["task_id"], sheet.cell(row=2, column=headers["任务ID"]).value)

    def test_legacy_netease_state_path_sets_provider_status(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            workbook_path = root / "tracker.xlsx"
            workbook = openpyxl.Workbook()
            workbook.active.title = "全部汇总"
            workbook.save(workbook_path)

            drafts_path = root / "drafts.jsonl"
            state_path = root / "netease-state.jsonl"
            draft = {
                "task_id": "c" * 24,
                "recipient": "teacher@example.edu",
                "subject": "测试主题",
                "body": "测试正文",
                "match_paragraph": "测试匹配段",
                "attachments": [],
                "teacher": {
                    "university": "测试大学",
                    "college": "信息学院",
                    "name": "陈测试",
                    "research_focus": "智能感知",
                    "faculty_url": "https://faculty.example.edu/teacher",
                    "match_score": "88",
                },
            }
            write_jsonl(drafts_path, [draft])
            write_jsonl(state_path, [{"task_id": draft["task_id"], "status": "saved"}])

            sync_workbook(workbook_path, drafts_path, state_path)
            result = openpyxl.load_workbook(workbook_path, data_only=True)
            sheet = result["全部汇总"]
            headers = {cell.value: cell.column for cell in sheet[1]}
            self.assertEqual("已存网易草稿", sheet.cell(row=2, column=headers["草稿状态"]).value)


if __name__ == "__main__":
    unittest.main()
